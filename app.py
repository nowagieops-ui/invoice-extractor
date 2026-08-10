import os, json, re, imaplib, email, shutil, base64, traceback
from email.header import decode_header
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google import genai
from google.genai import types as gtypes

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-in-production")

UPLOAD_FOLDER  = "uploads"
OUTPUT_FOLDER  = "output"
EXCEL_FILE     = os.path.join(OUTPUT_FOLDER, "INDA_Invoice_Tracker.xlsx")
PROCESSED_LOG  = os.path.join(OUTPUT_FOLDER, "processed_ids.json")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ── Style constants ────────────────────────────────────────────────────────
thin = Side(style="thin", color="000000")
STD_BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
APPROVED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
REJECTED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ROW_FILLS     = [
    PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"),
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
]

def sf(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)

# ── Excel: write a cell with style ────────────────────────────────────────
def wc(ws, r, c, val, font=None, fill=None, align=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:  cell.font  = font
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    cell.border = STD_BORDER
    return cell

# ── Excel: header setup ───────────────────────────────────────────────────
def _write_headers(ws):
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)

    # Row 1-2: Title
    ws.merge_cells("A1:Q1")
    ws["A1"] = "INDA PLATFORM"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws.merge_cells("A2:Q2")
    ws["A2"] = "INVOICES APPROVAL LINE TRACKING"
    ws["A2"].font = Font(name="Arial", bold=True, size=11)

    # Row 3: group headers
    groups = {7: "BUDGET HOLDER", 9: "AUDIT", 11: "COST CONTROL", 13: "GM FINANCE", 15: "SUBMIT TO FINANCE"}
    for col, label in groups.items():
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = CENTER

    # Row 4: column headers
    cols = ["S/N","SOFT COPY","VENDOR NAME","RECEIVED DATE\nBY FINANCE",
            "RECEIVED DATE\nBY COORDINATOR","PO/SO NUMBER",
            "DATE\nRECEIVE","DATE\nAPPROVE/REJECT",
            "DATE\nRECEIVE","DATE\nAPPROVE/REJECT",
            "DATE\nRECEIVE","DATE\nAPPROVE/REJECT",
            "DATE\nRECEIVE","DATE\nAPPROVE/REJECT",
            "NAME","DATE","REASON FOR REJECTION"]
    for i, label in enumerate(cols, 1):
        cell = ws.cell(row=4, column=i, value=label)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = CENTER; cell.border = STD_BORDER

    # Column widths
    widths = [8,10,32,14,14,28,12,20,12,20,12,20,12,20,14,12,32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "A5"

def get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["INVOICE"] if "INVOICE" in wb.sheetnames else wb.create_sheet("INVOICE")
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "INVOICE"
    _write_headers(ws)
    wb.save(EXCEL_FILE)
    return wb, ws

def next_sn(ws):
    """Auto-generate next S/N like 26-01, 26-02 …"""
    year = datetime.now().strftime("%y")
    max_num = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        sn = row[0]
        if sn and re.match(rf"^{year}-(\d+)$", str(sn)):
            max_num = max(max_num, int(re.findall(r"\d+$", str(sn))[0]))
    return f"{year}-{max_num+1:02d}"

def append_to_tracker(ws, wb, data: dict, row_index: int):
    """Write one extracted invoice record into the tracker."""
    r = row_index
    fill = ROW_FILLS[r % 2]

    approval_fill = lambda status: (
        APPROVED_FILL if status and "APPROVED" in status.upper()
        else REJECTED_FILL if status and "REJECT" in status.upper()
        else fill
    )
    approval_font = lambda status: (
        sf(bold=True, color="276749") if status and "APPROVED" in status.upper()
        else sf(bold=True, color="9C0006") if status and "REJECT" in status.upper()
        else sf()
    )

    wc(ws, r, 1,  data.get("sn"),               sf(bold=True), fill,  CENTER)
    wc(ws, r, 2,  None,                           sf(),          fill,  CENTER)
    wc(ws, r, 3,  data.get("vendor"),              sf(bold=True), fill,  LEFT)
    wc(ws, r, 4,  data.get("received_finance"),    sf(),          fill,  CENTER)
    wc(ws, r, 5,  data.get("received_coordinator"),sf(),          fill,  CENTER)
    wc(ws, r, 6,  data.get("po_so"),               sf(),          fill,  LEFT)

    # Budget Holder
    bh_s = data.get("budget_holder_status","")
    wc(ws, r, 7,  data.get("budget_holder_recv"),  sf(),          fill,  CENTER)
    wc(ws, r, 8,  f"{data.get('budget_holder_date','')} {bh_s}".strip(),
                                                    approval_font(bh_s), approval_fill(bh_s), CENTER)
    # Audit
    au_s = data.get("audit_status","")
    wc(ws, r, 9,  data.get("audit_recv"),           sf(),          fill,  CENTER)
    wc(ws, r, 10, f"{data.get('audit_date','')} {au_s}".strip(),
                                                    approval_font(au_s), approval_fill(au_s), CENTER)
    # Cost Control
    cc_s = data.get("cost_control_status","")
    wc(ws, r, 11, data.get("cost_control_recv"),    sf(),          fill,  CENTER)
    wc(ws, r, 12, f"{data.get('cost_control_date','')} {cc_s}".strip(),
                                                    approval_font(cc_s), approval_fill(cc_s), CENTER)
    # GM Finance
    gm_s = data.get("gm_finance_status","")
    wc(ws, r, 13, data.get("gm_finance_recv"),      sf(),          fill,  CENTER)
    wc(ws, r, 14, f"{data.get('gm_finance_date','')} {gm_s}".strip(),
                                                    approval_font(gm_s), approval_fill(gm_s), CENTER)
    # Submit to Finance
    wc(ws, r, 15, data.get("finance_name"),         sf(),          fill,  CENTER)
    wc(ws, r, 16, data.get("finance_date"),         sf(),          fill,  CENTER)
    # Rejection reason
    wc(ws, r, 17, data.get("rejection_reason"),     sf(),          fill,  LEFT)

    ws.row_dimensions[r].height = 45
    wb.save(EXCEL_FILE)

# ── AI extraction ──────────────────────────────────────────────────────────
def extract_with_ai(pdf_text: str) -> list:
    client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))

    prompt = """Extract invoice approval data from this PDF and return ONLY a JSON array.
No markdown, no explanation, just the raw JSON array.

Each element must have these exact keys (use null for missing values):
vendor, invoice_number, invoice_date, po_so, invoice_amount,
received_finance, invoice_period,
budget_holder_name, budget_holder_date, budget_holder_status,
audit_name, audit_date, audit_status,
cost_control_name, cost_control_date, cost_control_status,
gm_finance_name, gm_finance_date, gm_finance_status,
rejection_reason

Rules:
- Status fields must be exactly: APPROVED, REJECTED, or PENDING
- Dates in DD/MM/YYYY format only
- If signed and dated = APPROVED, if blank = PENDING, if crossed out = REJECTED
- For invoice_date use the last day of the service period if no date shown (e.g. 28/02/2026 for February 2026)
- Return an array always, even for one invoice

PDF TEXT:
""" + pdf_text[:5000]

    def clean_and_parse(raw):
        raw = raw.strip()
        raw = re.sub(r"^```json\s*|^```\s*|\s*```$", "", raw).strip()
        for pattern in [r'\[.*\]', r'\{.*\}']:
            m = re.search(pattern, raw, re.DOTALL)
            if m:
                try:
                    r = json.loads(m.group(0))
                    return [r] if isinstance(r, dict) else r
                except Exception:
                    pass
        try:
            r = json.loads(raw)
            return [r] if isinstance(r, dict) else r
        except Exception:
            return None

    # First attempt
    resp = client.models.generate_content(
        model="gemini-3.5-flash",
        contents=prompt,
        config=gtypes.GenerateContentConfig(temperature=0, max_output_tokens=2000)
    )
    result = clean_and_parse(resp.text)
    if result:
        return result

    # Second attempt — give Gemini an example to follow
    example = '[{"vendor":"TRAMATHY GLOBAL SERVICES LIMITED","invoice_number":"TGSL/BPL/KULA/02/2026","invoice_date":"28/02/2026","po_so":"BPL055C25-0104.05","invoice_amount":"3154775","received_finance":"02/03/2026","invoice_period":"FEBRUARY 2026","budget_holder_name":"SAMUEL ABEL-JUMBO","budget_holder_date":"04/08/2026","budget_holder_status":"APPROVED","audit_name":"BEN OKOH","audit_date":"06/08/2026","audit_status":"APPROVED","cost_control_name":null,"cost_control_date":null,"cost_control_status":"PENDING","gm_finance_name":null,"gm_finance_date":null,"gm_finance_status":"PENDING","rejection_reason":null}]'
    prompt2 = f"Return a JSON array like this example:\n{example}\n\nNow extract from this invoice text. Return ONLY the JSON array:\n{pdf_text[:3000]}"
    resp2 = client.models.generate_content(
        model="gemini-3.5-flash",
        contents=prompt2,
        config=gtypes.GenerateContentConfig(temperature=0, max_output_tokens=1500)
    )
    result2 = clean_and_parse(resp2.text)
    if result2:
        return result2

    raise ValueError(f"Could not parse Gemini response. Raw: {resp2.text[:300]}")


# ── PDF text extraction ────────────────────────────────────────────────────
def extract_pdf_text(filepath: str) -> str:
    text = ""
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    except Exception as e:
        text = f"[PDF read error: {e}]"
    return text

# ── Processed-email log ────────────────────────────────────────────────────
def load_processed():
    if os.path.exists(PROCESSED_LOG):
        with open(PROCESSED_LOG) as f:
            return set(json.load(f))
    return set()

def save_processed(ids: set):
    with open(PROCESSED_LOG, "w") as f:
        json.dump(list(ids), f)

# ── Outlook IMAP fetch ─────────────────────────────────────────────────────
def fetch_outlook_pdfs(email_addr, password, server, limit, folder="INBOX", sender_filter="", date_from=""):
    results = []
    debug = []
    mail = imaplib.IMAP4_SSL(server, 993)
    mail.login(email_addr, password)
    mail.select(folder)

    # Build IMAP search criteria
    search_parts = []
    if sender_filter.strip():
        # Use just the domain or local part to avoid case/display-name issues
        sf_clean = sender_filter.strip().lower()
        # Extract just the local part before @ for broader matching
        local_part = sf_clean.split("@")[0] if "@" in sf_clean else sf_clean
        search_parts.append(f'FROM "{local_part}"')
        debug.append(f"Sender filter: searching for '{local_part}' in From field")
    if date_from.strip():
        try:
            dt = datetime.strptime(date_from.strip(), "%Y-%m-%d")
            imap_date = dt.strftime("%d-%b-%Y")
            search_parts.append(f'SINCE {imap_date}')
            debug.append(f"Date filter: emails since {imap_date}")
        except ValueError:
            debug.append(f"⚠️ Invalid date format ignored: {date_from}")

    search_criteria = " ".join(search_parts) if search_parts else "ALL"
    debug.append(f"Search criteria: {search_criteria}")

    _, data = mail.search(None, search_criteria)
    all_ids = data[0].split()
    ids = all_ids[-limit:]
    debug.append(f"Emails matching search: {len(all_ids)} | Scanning last: {len(ids)}")

    processed = load_processed()
    debug.append(f"Already processed (skipping): {len(processed)}")

    skipped_processed = 0
    for msg_id in reversed(ids):
        uid = msg_id.decode()
        if uid in processed:
            skipped_processed += 1
            continue

        _, msg_data = mail.fetch(msg_id, "(RFC822)")
        raw_bytes = msg_data[0][1]
        msg = email.message_from_bytes(raw_bytes)

        subj_raw, enc = decode_header(msg.get("Subject",""))[0]
        subject = subj_raw.decode(enc or "utf-8") if isinstance(subj_raw, bytes) else str(subj_raw)
        sender  = msg.get("From","")
        date_str= msg.get("Date","")

        pdfs_found = []
        for part in msg.walk():
            fn = part.get_filename() or ""
            if fn.lower().endswith(".pdf"):
                fname_raw, fenc = decode_header(fn)[0]
                fname = fname_raw.decode(fenc or "utf-8") if isinstance(fname_raw, bytes) else str(fname_raw)
                save_path = os.path.join(UPLOAD_FOLDER, f"{uid}_{fname}")
                with open(save_path, "wb") as f_out:
                    f_out.write(part.get_payload(decode=True))
                pdfs_found.append({"path": save_path, "name": fname})

        if pdfs_found:
            debug.append(f"✅ From: {sender[:40]} | {subject[:40]} | PDFs: {[p['name'] for p in pdfs_found]}")
            results.append({
                "msg_id": uid,
                "subject": subject,
                "sender": sender,
                "date": date_str,
                "pdfs": pdfs_found
            })
        else:
            debug.append(f"⬜ {subject[:50]} — no PDF attachments")

    if skipped_processed:
        debug.append(f"ℹ️ Skipped {skipped_processed} already-processed emails")
    debug.append(f"📄 Emails with PDFs ready: {len(results)}")

    mail.logout()
    return results, processed, debug

# ── Core processing pipeline ───────────────────────────────────────────────
def process_emails(email_addr, password, server="imap.gmail.com", limit=20, sender_filter="", date_from=""):
    emails, processed, debug = fetch_outlook_pdfs(email_addr, password, server, limit, sender_filter=sender_filter, date_from=date_from)
    wb, ws = get_or_create_workbook()

    # Find next available row
    last_data_row = 4
    for row in ws.iter_rows(min_row=5):
        if row[0].value is not None:
            last_data_row = row[0].row
    next_row = last_data_row + 1
    row_index_counter = [next_row]

    summary = []
    new_processed = set()

    for em in emails:
        for pdf_info in em["pdfs"]:
            pdf_text = extract_pdf_text(pdf_info["path"])
            if len(pdf_text.strip()) < 50:
                summary.append({
                    "email": em["subject"],
                    "file": pdf_info["name"],
                    "status": "skipped",
                    "reason": "Could not extract text (scanned PDF or empty)"
                })
                continue

            try:
                records = extract_with_ai(pdf_text)
            except Exception as e:
                summary.append({
                    "email": em["subject"],
                    "file": pdf_info["name"],
                    "status": "error",
                    "reason": str(e)
                })
                continue

            for rec in records:
                sn = next_sn(ws)
                data = {
                    "sn": sn,
                    "vendor": rec.get("vendor"),
                    "received_finance": rec.get("received_finance") or rec.get("invoice_date"),
                    "received_coordinator": None,
                    "po_so": f"{rec.get('invoice_number','')} | {rec.get('po_so','')}",
                    "budget_holder_recv": rec.get("budget_holder_date"),
                    "budget_holder_date": rec.get("budget_holder_date"),
                    "budget_holder_status": rec.get("budget_holder_status","PENDING"),
                    "audit_recv": rec.get("audit_date"),
                    "audit_date": rec.get("audit_date"),
                    "audit_status": rec.get("audit_status","PENDING"),
                    "cost_control_recv": rec.get("cost_control_date"),
                    "cost_control_date": rec.get("cost_control_date"),
                    "cost_control_status": rec.get("cost_control_status","PENDING"),
                    "gm_finance_recv": rec.get("gm_finance_date"),
                    "gm_finance_date": rec.get("gm_finance_date"),
                    "gm_finance_status": rec.get("gm_finance_status","PENDING"),
                    "finance_name": None,
                    "finance_date": None,
                    "rejection_reason": rec.get("rejection_reason"),
                }
                append_to_tracker(ws, wb, data, row_index_counter[0])
                row_index_counter[0] += 1

                summary.append({
                    "email": em["subject"],
                    "file": pdf_info["name"],
                    "status": "ok",
                    "sn": sn,
                    "vendor": rec.get("vendor"),
                    "period": rec.get("invoice_period"),
                    "budget_holder": f"{rec.get('budget_holder_name','')} — {rec.get('budget_holder_status','')}",
                    "audit": f"{rec.get('audit_name','')} — {rec.get('audit_status','')}",
                    "rejection": rec.get("rejection_reason"),
                })

        new_processed.add(em["msg_id"])

    processed.update(new_processed)
    save_processed(processed)
    return summary, debug

# ─────────────────────────────────────────────────────────────────────────
# Flask routes
# ─────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    row_count = 0
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "INVOICE" in wb.sheetnames:
            ws = wb["INVOICE"]
            row_count = max(0, ws.max_row - 4)
    return render_template("index.html", row_count=row_count)


@app.route("/fetch-email", methods=["POST"])
def fetch_email():
    payload = request.get_json()
    email_addr    = payload.get("email","").strip()
    password      = payload.get("password","")
    server        = payload.get("server","imap.gmail.com")
    limit         = int(payload.get("limit", 20))
    sender_filter = payload.get("sender_filter","").strip()
    date_from     = payload.get("date_from","").strip()

    if not email_addr or not password:
        return jsonify({"error": "Email and password required"}), 400

    # Cap limit to avoid timeout on Render free tier (30s request limit)
    # Each email takes ~0.5s to fetch, so 50 is safe
    if limit > 200:
        limit = 200

    try:
        summary, debug = process_emails(email_addr, password, server, limit, sender_filter=sender_filter, date_from=date_from)
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    ok     = [s for s in summary if s["status"] == "ok"]
    errors = [s for s in summary if s["status"] != "ok"]
    return jsonify({"success": True, "processed": ok, "skipped": errors, "total": len(summary), "debug": debug})


@app.route("/upload", methods=["POST"])
def upload_pdf():
    """Manual single PDF upload — still useful for testing."""
    if "pdf" not in request.files:
        return jsonify({"error": "No file"}), 400
    file = request.files["pdf"]
    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "PDF files only"}), 400

    save_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(save_path)
    pdf_text = extract_pdf_text(save_path)

    if len(pdf_text.strip()) < 50:
        return jsonify({"error": "Could not read text from this PDF"}), 422

    try:
        records = extract_with_ai(pdf_text)
    except Exception as e:
        return jsonify({"error": f"AI extraction failed: {e}"}), 500

    wb, ws = get_or_create_workbook()
    last_data_row = 4
    for row in ws.iter_rows(min_row=5):
        if row[0].value is not None:
            last_data_row = row[0].row
    next_row = last_data_row + 1

    added = []
    for rec in records:
        sn = next_sn(ws)
        data = {
            "sn": sn,
            "vendor": rec.get("vendor"),
            "received_finance": rec.get("received_finance") or rec.get("invoice_date"),
            "received_coordinator": None,
            "po_so": f"{rec.get('invoice_number','')} | {rec.get('po_so','')}",
            "budget_holder_recv": rec.get("budget_holder_date"),
            "budget_holder_date": rec.get("budget_holder_date"),
            "budget_holder_status": rec.get("budget_holder_status","PENDING"),
            "audit_recv": rec.get("audit_date"),
            "audit_date": rec.get("audit_date"),
            "audit_status": rec.get("audit_status","PENDING"),
            "cost_control_recv": rec.get("cost_control_date"),
            "cost_control_date": rec.get("cost_control_date"),
            "cost_control_status": rec.get("cost_control_status","PENDING"),
            "gm_finance_recv": rec.get("gm_finance_date"),
            "gm_finance_date": rec.get("gm_finance_date"),
            "gm_finance_status": rec.get("gm_finance_status","PENDING"),
            "finance_name": None,
            "finance_date": None,
            "rejection_reason": rec.get("rejection_reason"),
        }
        append_to_tracker(ws, wb, data, next_row)
        next_row += 1
        added.append({"sn": sn, "vendor": rec.get("vendor"), "period": rec.get("invoice_period")})

    return jsonify({"success": True, "added": added, "count": len(added)})


@app.route("/preview")
def preview():
    rows = []
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "INVOICE" in wb.sheetnames:
            ws = wb["INVOICE"]
            for row in ws.iter_rows(min_row=5, values_only=True):
                if any(c is not None for c in row):
                    rows.append(row)
    headers = ["S/N","Soft Copy","Vendor Name","Recv by Finance","Recv by Coord","PO/SO",
               "BH Recv","BH Decision","Audit Recv","Audit Decision",
               "CC Recv","CC Decision","GM Recv","GM Decision",
               "Finance Name","Finance Date","Rejection Reason"]
    return render_template("preview.html", headers=headers, rows=rows)


@app.route("/download")
def download():
    if not os.path.exists(EXCEL_FILE):
        return redirect(url_for("index"))
    return send_file(EXCEL_FILE, as_attachment=True, download_name="INDA_Invoice_Tracker.xlsx")


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
